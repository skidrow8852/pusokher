import os
from pathlib import Path
from tkinter import Tk, Canvas, Entry,  PhotoImage,StringVar, messagebox, ttk
import openpyxl
from calculation import create_file, create_sheet_window, create_source_sheet, delete_default_sheet, fill_the_cells
from openpyxl import load_workbook
import argparse
from calculation import calculations

# Global variables
workbook = None
sheet_names = []
current_sheet_index = 0
notebook = None
experts = 5
max_percent = 10
min_percent = 3
step = 1
currentPercent = 0

def relative_to_assets(path: str) -> Path:
    current_path = Path(__file__).parent.absolute()  # Get the current directory
    assets_path = current_path / Path("assets/frame2")  # Set assets path
    return assets_path / Path(path)


# receive the file_path as a param from main.py file
def parse_arguments():
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="File path")

    return parser.parse_args()


args = parse_arguments()
file_path = args.file if args.file else "table.xlsx"



# Check for change on the Expert input
def saveExperts(*args):
    if entry_var_1.get() != "5":
       
            # Input is a valid number
            canvas.itemconfigure(image_8_active, state="normal")
            canvas.itemconfigure(image_8, state="hidden")
        
    else:
            canvas.itemconfigure(image_8_active, state="hidden")
            canvas.itemconfigure(image_8, state="normal")


# Set the Number of Experts
def NumofExperts():
    global experts
    try:
        value = int(entry_var_1.get())
        if value != experts:

            experts = value
            remove_excel_file()

            create_file(file_path)

            create_source_sheet(file_path, 1, experts, True)

            delete_default_sheet(file_path)

            LoadSheet()
        canvas.itemconfigure(image_8_active, state="hidden")
        canvas.itemconfigure(image_8, state="normal")
            
    except ValueError:
        # Input is not a valid number
        messagebox.showwarning("Error", "Пожалуйста, введите числовые значения.")


# Genearate Random data for the given sheet
def generateRand():
    current_tab = notebook.tab(notebook.select(), "text")
    numeric_value = ''.join(filter(str.isdigit, current_tab))
    fill_the_cells(file_path,numeric_value,experts)

    LoadSheet()


# Load the Excel Sheets
def LoadSheet():

    global workbook, sheet_names, notebook, current_sheet_index

    # Load the workbook
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames

    # Create the first sheet window
    notebook = ttk.Notebook(window)

    # Create a tab for each sheet and display the content
    for sheet_name in sheet_names:
        create_sheet_window(sheet_name, workbook, notebook, file_path, canvas,image_6_active)

    # Select the first sheet by default
    notebook.select(current_sheet_index)

    # Place the Notebook widget at the position and size of image_5
    
    x = 64
    y = 120.0
    width = 1421
    height = 474
    notebook.place(x=x, y=y, width=width, height=height)

    # Define the event handler function
    def check_selected_tab(event):
        global current_sheet_index
        current_tab = notebook.tab(notebook.select(), "text")
        current_tab_number = int(''.join(filter(str.isdigit, current_tab)))
        has_higher_tabs = False
    
        # Iterate over the tabs and check if there are higher number tabs
        for tab_id in notebook.tabs():
            tab_text = notebook.tab(tab_id, "text")
            tab_number = int(''.join(filter(str.isdigit, tab_text)))
            if tab_number > current_tab_number:
                has_higher_tabs = True
                break
        
        if current_tab.startswith("Исход") and not has_higher_tabs:
            if (currentPercent == 0 or currentPercent is None) or ( currentPercent > float(entry_var_2.get()) or currentPercent < float(entry_var_3.get())):
                canvas.itemconfigure(image_7_active, state="normal")
                canvas.itemconfigure(image_7, state="hidden")
            
            else:
                canvas.itemconfigure(image_7_active, state="hidden")
                canvas.itemconfigure(image_7, state="normal")
        else:
            canvas.itemconfigure(image_7_active, state="hidden")
            canvas.itemconfigure(image_7, state="normal")
        
        current_sheet_index = notebook.index(notebook.select())
    # Bind the event handler to the NotebookTabChanged event
    notebook.bind("<<NotebookTabChanged>>", check_selected_tab)



# Handle Next Step (Display Next sheet)
def handle_next(notebook):
    global current_sheet_index
    current_sheet_index += 1
    if current_sheet_index >= len(sheet_names):
        current_sheet_index = 0

    notebook.select(current_sheet_index)



# Handle Previous Step (Display Previous sheet)
def handle_previous(notebook):
    global current_sheet_index
    current_sheet_index -= 1
    if current_sheet_index < 0:
        current_sheet_index = len(sheet_names) - 1

    notebook.select(current_sheet_index)


# Calculation for Each Step
def NextStep(max_possible_percent,min_possible_percent):
    global currentPercent, step
    
    if (currentPercent == 0 or currentPercent is None) or ( currentPercent > max_possible_percent or currentPercent < min_possible_percent):
        if step < 6:
            currentPercent = calculations(file_path, step,1000,experts)
    




# Check if the data on the sheet are correct and valid
def check_data():
    if(step < 6):
        wb = openpyxl.load_workbook(file_path)

        sheet_name = f'Исходные данные {step} шага'
        sheet = wb[sheet_name]

        for row_num in range(2, experts + 2):
            for column_num in range(2, 5):
                cell = sheet.cell(row=row_num, column=column_num)
                if cell.value is None:
                    messagebox.showwarning("Error", f'в ячейке отсутствуют данные ({cell.coordinate})')
                    wb.close()
                    return False
                elif not isinstance(cell.value, (float, int)):
                    messagebox.showwarning("Error", f'Исходные данные {step} шага не верны или не правильном формате')
                
                    wb.close()
                    return False

        wb.close()
        return True


# Function responsible for Sheet Creation for each step and input validation
def validate_inputs():
    global step, current_sheet_index, currentPercent
    all_numeric = all(var.get().isdigit() for var in input_vars)
    if not all_numeric:
        messagebox.showwarning("Error", "Пожалуйста, введите числовые значения.")
    else:
        
        entry_1.config(state='disabled', disabledbackground='#0F0D1D', disabledforeground='#878787')
        entry_2.config(state='disabled', disabledbackground='#0F0D1D', disabledforeground='#878787')
        entry_3.config(state='disabled', disabledbackground='#0F0D1D', disabledforeground='#878787')

        if check_data():
            if step <=5:
                NextStep(float(entry_var_2.get()), float(entry_var_3.get()))   

                if (currentPercent == 0 or currentPercent is None) or ( currentPercent > float(entry_var_2.get()) or currentPercent < float(entry_var_3.get())):
                    step+=1
                    if step <6:
                        create_source_sheet(file_path, step,experts,True)
                        current_sheet_index += 1
            LoadSheet()

            canvas.itemconfigure(image_9_active, state="normal")
            canvas.itemconfigure(image_9, state="hidden")
        
    if (currentPercent != 0 and currentPercent is not None) and currentPercent < float(entry_var_2.get()) and currentPercent > float(entry_var_3.get()):    
        handle_next(notebook)

# Delete an excel file after Reset
def remove_excel_file():
    if os.path.exists(file_path):
        os.remove(file_path)


# Reset all the data and remove the sheets
def resetAll():
    global current_sheet_index, step, currentPercent
    current_sheet_index = 0
    step = 1
    currentPercent = 0
    remove_excel_file()

    create_file(file_path)

    create_source_sheet(file_path, 1, experts, True)

    delete_default_sheet(file_path)

    LoadSheet()

    entry_1.config(state='normal')
    entry_2.config(state='normal' )
    entry_3.config(state='normal')

    entry_var_1.set(f"{experts}")
    entry_var_2.set("10")
    entry_var_3.set("3")


window = Tk()
window.title("ПУЗ-ОРХ")
style = ttk.Style(window)
window.tk.call("source", "assets/forest-dark.tcl")
style.theme_use("forest-dark")


window.geometry("1541x832")
window.configure(bg = "#06010F")



canvas = Canvas(
    window,
    bg = "#06010F",
    height = 832,
    width = 1541,
    bd = 0,
    highlightthickness = 0,
    relief = "ridge"
)

canvas.place(x = 0, y = 0)
image_image_1 = PhotoImage(
    file=relative_to_assets("image_1.png"))
image_1 = canvas.create_image(
    771.0,
    669.0,
    image=image_image_1
)

image_image_2 = PhotoImage(
    file=relative_to_assets("image_2.png"))
image_2 = canvas.create_image(
    1281.0,
    768.0,
    image=image_image_2
)

image_image_3 = PhotoImage(
    file=relative_to_assets("image_3.png"))
image_3 = canvas.create_image(
    1416.0,
    773.0,
    image=image_image_3
)

image_image_4 = PhotoImage(
    file=relative_to_assets("image_4.png"))
image_4 = canvas.create_image(
    932.0,
    372.0,
    image=image_image_4
)

image_image_5 = PhotoImage(
    file=relative_to_assets("image_5.png"))
image_5 = canvas.create_image(
    775.0,
    659.0,
    image=image_image_5
)

image_image_6 = PhotoImage(
    file=relative_to_assets("image_6.png"))
image_6 = canvas.create_image(
    1346.0,
    660.0,
    image=image_image_6
)

image_image_6_active = PhotoImage(
    file=relative_to_assets("image_6_active.png"))
image_6_active = canvas.create_image(
    1346.0,
    660.0,
    image=image_image_6_active,
    
)

image_image_7 = PhotoImage(
    file=relative_to_assets("image_7.png"))
image_7 = canvas.create_image(
    1168.0,
    660.0,
    image=image_image_7
)

image_image_7_active = PhotoImage(
    file=relative_to_assets("image_7_active.png"))
image_7_active = canvas.create_image(
    1168.0,
    660.0,
    image=image_image_7_active
)

entry_image_1 = PhotoImage(
    file=relative_to_assets("entry_1.png"))
entry_bg_1 = canvas.create_image(
    181.5,
    83.5,
    image=entry_image_1
)
entry_var_1 = StringVar()
entry_var_1.set("5")
entry_1 = Entry( bd=0, bg="#0F0D1D",  fg='white', font=("Arial", 18), highlightthickness=0,highlightcolor='white'
                ,highlightbackground='white',textvariable=entry_var_1)
entry_1.place(
    x=106.0,
    y=56.0,
    width=151.0,
    height=53.0
)

canvas.create_text(
    91.0,
    23.0,
    anchor="nw",
    text="Кол-во экспертов",
    fill="#FFFFFF",
    font=("PlusJakartaSansRoman SemiBold", 17 * -1)
)

image_image_8 = PhotoImage(
    file=relative_to_assets("image_8.png"))
image_8 = canvas.create_image(
    388.0,
    83.0,
    image=image_image_8
)


image_image_8_active = PhotoImage(
    file=relative_to_assets("image_8_active.png"))
image_8_active = canvas.create_image(
    388.0,
    83.0,
    image=image_image_8_active,
    state='hidden'
)


canvas.create_text(
    524.0,
    23.0,
    anchor="nw",
    text="Максимальный  процент",
    fill="#FFFFFF",
    font=("PlusJakartaSansRoman SemiBold", 17 * -1)
)

entry_image_2 = PhotoImage(
    file=relative_to_assets("entry_2.png"))
entry_bg_2 = canvas.create_image(
    614.5,
    83.5,
    image=entry_image_2
)

entry_var_2 = StringVar()
entry_var_2.set("10")
entry_2 = Entry( bd=0, bg="#0F0D1D",  fg='white', font=("Arial", 18), highlightthickness=0,highlightcolor='white',
                highlightbackground='white' , textvariable=entry_var_2)


entry_2.place(
    x=539.0,
    y=56.0,
    width=151.0,
    height=53.0
)

entry_image_3 = PhotoImage(
    file=relative_to_assets("entry_3.png"))
entry_bg_3 = canvas.create_image(
    886.5,
    83.5,
    image=entry_image_3
)
entry_var_3 = StringVar()
entry_var_3.set("3")
entry_3 = Entry( bd=0, bg="#0F0D1D",  fg='white', font=("Arial", 18), highlightthickness=0,highlightcolor='white',
                highlightbackground='white', textvariable=entry_var_3)
entry_3.place(
    x=811.0,
    y=56.0,
    width=151.0,
    height=53.0
)

canvas.create_text(
    796.0,
    23.0,
    anchor="nw",
    text="Минимальный  процент",
    fill="#FFFFFF",
    font=("PlusJakartaSansRoman SemiBold", 17 * -1)
)

image_image_9 = PhotoImage(
    file=relative_to_assets("image_9.png"))
image_9 = canvas.create_image(
    1269.0,
    72.0,
    image=image_image_9
)

image_image_9_active = PhotoImage(
    file=relative_to_assets("image_9_active.png"))
image_9_active = canvas.create_image(
    1269.0,
    72.0,
    image=image_image_9_active,
    state='hidden'
)



input_vars = [entry_var_1, entry_var_2, entry_var_3]  # List of input fields
entry_var_1.trace_add("write", saveExperts)  # Bind trace to each entry

saveExperts()
LoadSheet()
canvas.tag_raise(image_3)
canvas.tag_raise(image_2)
canvas.tag_bind(image_3, "<Button-1>", lambda event: validate_inputs())
canvas.tag_bind(image_8_active, "<Button-1>", lambda event: NumofExperts())
canvas.tag_bind(image_7_active, "<Button-1>", lambda event: generateRand())
canvas.tag_bind(image_9_active, "<Button-1>", lambda event: resetAll())
canvas.tag_bind(image_2, "<Button-1>", lambda event: handle_previous(notebook))


window.resizable(False, False)
window.mainloop()
