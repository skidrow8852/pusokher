import copy
import random
from statistics import mean
from math import sqrt
from io import BytesIO
import scipy.stats as stats
import tkinter as tk
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl.drawing.image import Image
# Global variables
sheet_names = []
current_sheet_index = 0
iteration_number = 1000
# DIR WHERE IMAGES SAVED TO
# WHEN RUNNING MAKE SURE ITS FILLED WITH DIRECTORY OF FILE
source_dir = ""


def format_number(value):
    if isinstance(value, int):
        return str(value)
    elif isinstance(value, float):
        if value.is_integer():
            return "{:.0f}".format(value)
        else:
            return "{:.2f}".format(value)
    else:
        return str(value)




def create_sheet_window(sheet_name, workbook, notebook,file=None,canvas=None, image=None):
    # Read data from the sheet
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        formatted_row = [format_number(cell) if isinstance(cell, (int, float)) else cell for cell in row]
        data.append(formatted_row)

    # Create a frame for the sheet's content
    sheet_frame = tk.Frame(notebook)
    notebook.add(sheet_frame, text=sheet_name)

    # Create a canvas for the sheet's content
    sheet_canvas = tk.Canvas(sheet_frame)
    sheet_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Create a scrollbar for vertical scrolling
    sheet_scrollbar = tk.Scrollbar(sheet_frame, orient=tk.VERTICAL, command=sheet_canvas.yview)
    sheet_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    sheet_canvas.configure(yscrollcommand=sheet_scrollbar.set)

    # Create a frame inside the canvas for the sheet data
    sheet_data_frame = tk.Frame(sheet_canvas)
    sheet_canvas.create_window((0, 0), window=sheet_data_frame, anchor=tk.NW)

    # Create a label for each cell and display the data with increased font size
    font_size = 11  # Choose your desired font size
    for row_index, row_data in enumerate(data):
        for col_index, cell_data in enumerate(row_data):
            label = tk.Label(sheet_data_frame, text=cell_data, font=("TkDefaultFont", font_size))
            label.grid(row=row_index, column=col_index, padx=5, pady=5)

    # Configure the canvas to adjust the scroll region based on the frame size
    def configure_canvas(event):
        sheet_canvas.configure(scrollregion=sheet_canvas.bbox("all"))

    sheet_data_frame.bind("<Configure>", configure_canvas)

    # Add mouse wheel scrolling functionality to the canvas
    def scroll_canvas(event):
        sheet_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    sheet_canvas.bind_all("<MouseWheel>", scroll_canvas)

    # Create an entry widget for specified cells to make them editable
    entry_fields = {
        "B": (2, sheet.max_row),
        "C": (2, sheet.max_row),
        "D": (2, sheet.max_row),
        "F": (2, sheet.max_row),
        "G": (2, sheet.max_row),
        "H": (2, sheet.max_row),
    }
    entries = []
    entry_values = {}  # Keep track of original entry values

    def entry_change(event, entry):
        entry_values[entry] = entry.get()  # Store the current entry value
        entry_value = entry.get()
        col_index = entry.grid_info()["column"] - 1
        row_index = entry.grid_info()["row"] - 2 + entry_fields[openpyxl.utils.get_column_letter(col_index + 2)][0]
                
        try:  
            sheet.cell(row=row_index + 1, column=col_index + 2).value = int(entry_value)
            workbook.save(file)
        except ValueError:
            print("")

    for col_letter, (start_row, end_row) in entry_fields.items():
        col_index = openpyxl.utils.column_index_from_string(col_letter) - 1

        # Check if the column index is within bounds
        if col_index >= len(data[0]):
            continue
        
        if sheet_name.startswith("Исход"):
            for row_index in range(start_row - 1, end_row):
                cell_data = data[row_index][col_index]
                entry = tk.Entry(sheet_data_frame, width=20 if col_letter in ['F', 'G', 'H'] else 10)
                entry.insert(tk.END, str(cell_data if cell_data else ""))
                entry.grid(row=row_index, column=col_index, padx=5, pady=5)
                entry.bind("<Key>", lambda event, entry=entry: entry_change(event, entry))  # Track entry changes
                entries.append(entry)
                entry_values[entry] = entry.get()  # Store the original entry value

    # Configure the canvas to adjust the scroll region based on the frame size
    def configure_canvas(event):
        sheet_canvas.configure(scrollregion=sheet_canvas.bbox("all"))

    sheet_data_frame.bind("<Configure>", configure_canvas)

    # Add mouse wheel scrolling functionality to the canvas
    def scroll_canvas(event):
        sheet_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    sheet_canvas.bind_all("<MouseWheel>", scroll_canvas)

    # Function to save the changed data to the sheet
    def save_changed_data():
        changed_fields = []
        
        for entry in entries:
            if entry_values[entry] != entry.get():  # Compare current entry value with original value
                changed_fields.append(entry)
        
        
        for entry in changed_fields:
            entry_value = entry.get()
            
            col_index = entry.grid_info()["column"] - 1
            row_index = entry.grid_info()["row"] - 2 + entry_fields[openpyxl.utils.get_column_letter(col_index + 2)][0]
            
            # Convert the entry value to a float
            col_letter = openpyxl.utils.get_column_letter(col_index + 2)
            if col_letter in ["B", "C", "D"]:
                try:
                    entry_value = float(entry_value)
                except ValueError:
                    print("Error: Invalid number entered")
                    continue
            
            sheet.cell(row=row_index + 1, column=col_index + 2).value = entry_value
        
        # Save the workbook back to the file
        workbook.save(file)


    # Create a save button for saving the changed data to the sheet
    canvas.tag_bind(image, "<Button-1>", lambda event: save_changed_data())

    return sheet_frame


# Creates a sheet with the graphs images for each step
def create_image_sheet(filepath, step, number_of_experts):
    # Create a new worksheet only if it doesn't already exist.
    workpath = image_path = os.path.dirname(filepath) + f"/graphs.xlsx"
    if not os.path.exists(workpath):
        create_file(workpath)
    wb = openpyxl.load_workbook(workpath)
    sheet_name = f"Graphs {step} шаг"
    worksheet = None
    if sheet_name in wb.sheetnames:
        worksheet = wb[sheet_name]
    else:
        worksheet = wb.create_sheet(title=sheet_name)

    for row_num in range(1, number_of_experts + 1):
        # Insert an image.
        image_path = os.path.dirname(filepath) + f"/распределение_{step}_шаг_{row_num}_эксперт.png"

        img = Image(image_path)
        # Resize the image to 450x400 pixels.
        img.width = 450
        img.height = 400
        column_letter = openpyxl.utils.get_column_letter(row_num)
        worksheet.add_image(img, anchor=f"{column_letter}3")

        worksheet.cell(row=1, column=row_num, value=f"{step}_шаг - {row_num}_эксперт")
        # Set the column width based on the image width.
        column_width = img.width / 8
        worksheet.column_dimensions[column_letter].width = column_width

    # Save the modified workbook.
    try:
        wb.save(workpath)
        delete_default_sheet(workpath)
    except:
        # File is closed, open it and save the workbook
        wb = openpyxl.load_workbook(workpath)
        wb.save(workpath)
        delete_default_sheet(workpath)
    

# function for development assistance
def fill_the_cells(filepath, step, number_of_experts=5):
    wb = openpyxl.load_workbook(filepath)

    sheet_name = f'Исходные данные {step} шага'
    sheet = wb[sheet_name]

    # Generate random data and fill the table
    for row_num in range(2, number_of_experts + 2):
        max_value = random.randint(8, 10)
        sheet.cell(row=row_num, column=4, value=max_value)

        avg_value = random.randint(5, 7)
        sheet.cell(row=row_num, column=3, value=avg_value)

        min_value = random.randint(1, 4)
        sheet.cell(row=row_num, column=2, value=min_value)

    wb.save(filepath)


def create_file(filepath):
    wb = openpyxl.Workbook()

    wb.save(filepath)
    return wb


def distribution(filepath, step, source_dir, number_of_experts):
    create_source_sheet(filepath, step, number_of_experts, needed=False, distr=True)
    wb = openpyxl.load_workbook(filepath)

    
    source_sheet_name = f'Исходные данные {step} шага'
    source_sheet = wb[source_sheet_name]

    distribution_sheet_name = f'Распределенные данные {step} шага'
    distribution_sheet = wb[distribution_sheet_name]

    indexes = {"минимально": 2, "среднее": 3, "максимально": 4}

    min_column_values = []
    for row in source_sheet.iter_rows(min_row=1, values_only=True):
        value = row[indexes["минимально"] - 1]
        if isinstance(value, int):
            min_column_values.append(value)

    max_column_values = []
    for row in source_sheet.iter_rows(min_row=1, values_only=True):
        value = row[indexes["максимально"] - 1]
        if isinstance(value, int):
            max_column_values.append(value)

    most_likely_column_values = []
    for row in source_sheet.iter_rows(min_row=1, values_only=True):
        value = row[indexes["среднее"] - 1]
        if isinstance(value, int):
            most_likely_column_values.append(value)

    # Создаем массив для хранения случайных оценок
    random_scores = np.zeros((len(min_column_values), iteration_number))

    for i in range(number_of_experts):
        scale = max_column_values[i] - min_column_values[i]
        c = (most_likely_column_values[i] - min_column_values[i]) / scale
        loc = min_column_values[i]

        distribution = stats.triang(loc=loc, c=c, scale=scale)
        random_scores[i] = distribution.rvs(size=iteration_number)

    # запись статистических показателей для каждого эксперта
    for i in range(number_of_experts):
        distribution_sheet.cell(column=2, row=i + 2, value=np.min(random_scores[i]))
        distribution_sheet.cell(column=3, row=i + 2, value=np.mean(random_scores[i]))
        distribution_sheet.cell(column=4, row=i + 2, value=np.max(random_scores[i]))

        # Save plot image to a byte stream
        image_stream = BytesIO()
        plt.hist(random_scores[i], bins='auto')
        plt.title('Распределение для Эксперта {}'.format(i + 1))
        plt.xlabel('Значение')
        plt.ylabel('Частота')
        plt.savefig(image_stream, format='png')

        # Rewind the byte stream position to the beginning
        image_stream.seek(0)

        # Save plot image to a file in the specified image directory
        image_path = os.path.join(source_dir, f'распределение_{step}_шаг_{i + 1}_эксперт.png')
        plt.savefig(image_path, format='png')
        plt.close()

        
        

    wb.save(filepath)
    




def create_source_sheet(filepath, step, number_of_experts=5, needed=False, distr=None):
    wb = openpyxl.load_workbook(filepath)

    if distr is None:
        sheet_name = f'Исходные данные {step} шага'
    else:
        sheet_name = f'Распределенные данные {step} шага'

    sheet = wb.create_sheet(title=sheet_name)

    column_names = ['минимально', 'наиболее вероятно', 'максимально']

    # generating rows with experts
    for row_num in range(1, number_of_experts + 1):
        cell_value = f'E{row_num}'
        sheet.cell(row=row_num + 1, column=1, value=cell_value)

    # generating columns with values
    for col_num, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=col_num + 1, value=column_name)

        # generating feedback columns
        if needed:
            if distr is None:
                for reason in range(len(column_names)):
                    sheet.cell(row=1, column=reason + len(column_names) + 3, value=f"Объяснение {column_names[reason]}")

    wb.save(filepath)


def create_calculation_sheet(filepath, step, number_of_experts=5):
    wb = openpyxl.load_workbook(filepath)

    sheet_name = f'Вычисления {step} шага'
    sheet = wb.create_sheet(title=sheet_name)

    columns_names = ['Число итераций', 'Среднее оценок экспертов',
                     'Дисперсия', 'Среднеквадр. отклонение', 'Коэф. вариации', 'Асимметрия', ]

    # generating mean row
    sheet.cell(row=2, column=1, value="Среднее каждого столбца")

    # generating rows with experts
    for row_num in range(1, number_of_experts + 1):
        cell_value = f'E{row_num}'
        sheet.cell(row=row_num + 2, column=1, value=cell_value)

    # generating columns with values
    for col_num, column_name in enumerate(columns_names, start=1):
        sheet.cell(row=1, column=col_num + 1, value=column_name)

    wb.save(filepath)

    return wb, sheet


def calculations(filepath, step, total_number_of_experts=1000, number_of_experts=5, source="."):

    distribution(filepath, step, source, number_of_experts)
    
    wb, calculation_sheet = create_calculation_sheet(filepath, step, number_of_experts)
     
    create_image_sheet(filepath,step,number_of_experts)
    
    source_sheet_name = f'Распределенные данные {step} шага'
    source_sheet = wb[source_sheet_name]

    

    iteration_number = total_number_of_experts

    calculation_columns_names = ['Число итераций', 'Среднее оценок экспертов',
                                 'Дисперсия', 'Среднеквадр. отклонение', 'Коэф. вариации', 'Асимметрия', ]

    
    

    # заполнение поля число итераций
    for i in range(number_of_experts):
        calculation_sheet.cell(row=i + 3, column=2, value=iteration_number)

    def find_column_index_by_name(column_name):

        # find the column index based on the column name
        column_index = None
        for cell in calculation_sheet[1]:
            if cell.value == column_name:
                column_index = cell.column_letter
                return column_index

    def convert_column_letter_to_number(column_letter):
        column_number = 0
        power = 1
        for char in reversed(column_letter):
            char_value = ord(char.upper()) - ord('A') + 1
            column_number += char_value * power
            power *= 26
        return column_number

    # вычисление среднеарифметического
    def arithmetic_mean(column_name=None, row=None):

        if column_name:
            column_values = []
            column_index = find_column_index_by_name(column_name)

            if column_index:
                for cell in calculation_sheet[column_index]:
                    if isinstance(cell.value, (int, float)):
                        column_values.append(cell.value)

            if column_values:
                result = mean(column_values)
                return result, column_index

        if row:
            row_values = []
            for column in range(2, 5):
                cell = source_sheet.cell(row=row, column=column).value
                if isinstance(cell, (int, float)):
                    row_values.append(cell)

            result = mean(row_values)
            return result

        return []

    # вычисление дисперсии
    def variance(expert):
        expert_row_source_sheet = expert + 1
        min_rate = source_sheet.cell(column=2, row=expert_row_source_sheet).value
        avg_rate = source_sheet.cell(column=3, row=expert_row_source_sheet).value
        max_rate = source_sheet.cell(column=4, row=expert_row_source_sheet).value

        expert_row_calculation_sheet = expert + 2
        rates_mean = calculation_sheet.cell(column=3, row=expert_row_calculation_sheet).value

        var = ((min_rate - rates_mean) ** 2 + (avg_rate - rates_mean) ** 2 + (max_rate - rates_mean) ** 2) / 3

        return var

    # вычисление среднеквадратического отклонения
    def deviation(expert):
        expert_row_source_sheet = expert + 1
        expert_row_calculation_sheet = expert + 2

        dev = sqrt(calculation_sheet.cell(column=4, row=expert_row_calculation_sheet).value)

        return dev

    # вычисление асимметрии
    def asymmetry(expert):
        expert_row_source_sheet = expert + 1
        expert_row_calculation_sheet = expert + 2

        rates_mean = calculation_sheet.cell(column=3, row=expert_row_calculation_sheet).value
        max_rate = source_sheet.cell(column=4, row=expert_row_source_sheet).value

        dev = sqrt(calculation_sheet.cell(column=4, row=expert_row_calculation_sheet).value)

        asym = (rates_mean - max_rate) / dev

        return asym

    # вычисление коэффициента вариации
    def variation_coefficient(expert):
        expert_row_calculation_sheet = expert + 2
        deviation_number = calculation_sheet.cell(column=5, row=expert_row_calculation_sheet).value
        rates_mean = calculation_sheet.cell(column=3, row=expert_row_calculation_sheet).value

        coefficient = deviation_number / rates_mean

        return coefficient

    def percent_calculation():
        previous_step = step - 1
        previous_calculation_sheet_name = f'Вычисления {previous_step} шага'

        previous_calculation_sheet = wb[previous_calculation_sheet_name]

        prev_variation = previous_calculation_sheet.cell(row=2, column=6).value
        current_variation = calculation_sheet.cell(row=2, column=6).value

        result = abs(prev_variation - current_variation) * 100 / prev_variation

        return result

    # 1: заполнение ячеек столбца среднее оценок экспертов
    for expert_row in range(2, number_of_experts + 2):
        column_letter = find_column_index_by_name('Среднее оценок экспертов')
        column = convert_column_letter_to_number(column_letter)

        value = arithmetic_mean(row=expert_row)
        calculation_sheet.cell(row=expert_row + 1, column=column, value=value)

    # 2: заполнение ячеек с дисперсией
    for i in range(1, number_of_experts + 1):
        calculation_sheet.cell(column=4, row=i + 2, value=variance(i))

    # 3: заполнение ячеек среднееквадр. отклонение
    for i in range(1, number_of_experts + 1):
        calculation_sheet.cell(column=5, row=i + 2, value=deviation(i))

    # 4: заполнение ячеек коэффициент вариации
    for i in range(1, number_of_experts + 1):
        calculation_sheet.cell(column=6, row=i + 2, value=variation_coefficient(i))

    # 5: заполнение ячеек асимметрии
    for i in range(1, number_of_experts + 1):
        calculation_sheet.cell(column=7, row=i + 2, value=asymmetry(i))

    # 6: среднеарифметическое для каждого слолбца
    for column in calculation_columns_names:
        result, column_index = arithmetic_mean(column_name=column)
        column_index = convert_column_letter_to_number(column_index)
        if result:
            calculation_sheet.cell(row=2, column=column_index, value=result)

    wb.save(filepath)

    if step != 1:
        percent = percent_calculation()
        # добавление столбца Процент оценки спроса
        calculation_sheet.cell(row=1, column=9, value="Процент оценки спроса")
        calculation_sheet.cell(row=2, column=9, value=percent)
        wb.save(filepath)

        return percent

    
# remove default sheet (empty sheet created by openyxl)
def delete_default_sheet(filepath):
    wb = openpyxl.load_workbook(filepath)
    default_sheet_name = 'Sheet'
   # Check if the default sheet exists
    if default_sheet_name in wb.sheetnames:
        default_sheet = wb[default_sheet_name]
        wb.remove(default_sheet)  # Remove the default sheet
        wb.save(filepath)
